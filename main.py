import speech_recognition as sr
import openpyxl
from urllib.parse import quote
import os

# Nome do arquivo Excel
ARQUIVO = "lista_compras.xlsx"

# Inicializa a planilha se não existir
def inicializar_planilha():
    if not os.path.exists(ARQUIVO):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compras"
        ws.append(["Produto"])
        wb.save(ARQUIVO)

def adicionar_produto(produto):
    wb = openpyxl.load_workbook(ARQUIVO)
    ws = wb.active
    ws.append([produto])
    wb.save(ARQUIVO)
    print(f"[AGENTE] Produto '{produto}' adicionado à lista.")

def remover_produto(produto):
    wb = openpyxl.load_workbook(ARQUIVO)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value and produto.lower() in row[0].value.lower():
            ws.delete_rows(row[0].row, 1)
            wb.save(ARQUIVO)
            print(f"[AGENTE] Produto '{produto}' removido da lista.")
            return
    print(f"[AGENTE] Produto '{produto}' não encontrado.")

def listar_produtos():
    wb = openpyxl.load_workbook(ARQUIVO)
    ws = wb.active
    produtos = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    if produtos:
        print("[AGENTE] Sua lista contém:", ", ".join(produtos))
    else:
        print("[AGENTE] Sua lista está vazia.")

def enviar_lista():
    wb = openpyxl.load_workbook(ARQUIVO)
    ws = wb.active
    produtos = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    if produtos:
        mensagem = "Lista de compras: " + ", ".join(produtos)
        link = f"https://wa.me/?text={quote(mensagem)}"
        print("[AGENTE] Envie sua lista pelo link abaixo:")
        print(link)
    else:
        print("[AGENTE] Sua lista está vazia, nada para enviar.")

def interpretar_comando(comando):
    comando = comando.lower()
    if "adicionar" in comando:
        produto = comando.replace("adicionar", "").strip()
        if produto:
            adicionar_produto(produto)
        else:
            print("[AGENTE] Diga o produto a ser adicionado.")
    elif "remover" in comando or "tirar" in comando:
        produto = comando.replace("remover", "").replace("tirar", "").strip()
        if produto:
            remover_produto(produto)
        else:
            print("[AGENTE] Diga o produto a ser removido.")
    elif "listar" in comando or "mostrar" in comando:
        listar_produtos()
    elif "enviar" in comando:
        enviar_lista()
    elif "sair" in comando or "parar" in comando:
        print("[AGENTE] Encerrando...")
        return False
    else:
        print("[AGENTE] Não entendi o comando.")
    return True

def ouvir_comando():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        print("\n[AGENTE] Fale seu comando:")
        audio = r.listen(source)
        try:
            comando = r.recognize_google(audio, language="pt-BR")
            print(f"[Você]: {comando}")
            return comando
        except sr.UnknownValueError:
            print("[AGENTE] Não entendi o que você disse.")
        except sr.RequestError:
            print("[AGENTE] Erro no serviço de reconhecimento.")
    return ""

def main():
    inicializar_planilha()
    print("=== Agente de Compras com MCP (versão simplificada) ===")
    ativo = True
    while ativo:
        comando = ouvir_comando()
        if comando:
            ativo = interpretar_comando(comando)

if __name__ == "__main__":
    main()
