
import os
import json
import time
import openpyxl
from urllib.parse import quote
import webbrowser

# === (Opcional) STT stack igual ao seu: Vosk + sounddevice ===
# Mantemos como está, mas o foco aqui é a LLM para interpretar linguagem natural.
try:
    from vosk import Model, KaldiRecognizer
    import sounddevice as sd
    VOSK_OK = True
except Exception:
    VOSK_OK = False

# === LLM Providers ===
# Suporta OpenAI (OPENAI_API_KEY) ou Ollama local (http://localhost:11434, OLLAMA_MODEL)
import requests

ARQUIVO = "lista_compras.xlsx"

# --------------------- UTIL: Planilha ---------------------
def inicializar_planilha():
    if not os.path.exists(ARQUIVO):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compras"
        ws.append(["Produto"])
        wb.save(ARQUIVO)

def adicionar_produto(produto):
    wb = openpyxl.load_workbook(ARQUIVO)
    ws = wb.active
    ws.append([produto])
    wb.save(ARQUIVO)
    print(f"[AGENTE] Produto '{produto}' adicionado.")

def remover_produto(produto):
    wb = openpyxl.load_workbook(ARQUIVO)
    ws = wb.active
    # Remove linhas que contenham o termo (case-insensitive)
    removed = False
    for row in list(ws.iter_rows(min_row=2, values_only=False)):
        if row[0].value and produto.lower() in str(row[0].value).lower():
            ws.delete_rows(row[0].row, 1)
            removed = True
    wb.save(ARQUIVO)
    if removed:
        print(f"[AGENTE] Removi itens que batiam com '{produto}'.")
    else:
        print(f"[AGENTE] Não encontrei '{produto}'.")

def listar_produtos(retornar=False):
    wb = openpyxl.load_workbook(ARQUIVO)
    ws = wb.active
    produtos = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    if retornar:
        return produtos
    if produtos:
        print("[AGENTE] Lista:", ", ".join(produtos))
    else:
        print("[AGENTE] Sua lista está vazia.")

def enviar_lista(abrir_navegador=True):
    produtos = listar_produtos(retornar=True)
    if produtos:
        mensagem = "Lista de compras: " + ", ".join(produtos)
        link = f"https://wa.me/?text={quote(mensagem)}"
        print("[AGENTE] Link para enviar no WhatsApp:")
        print(link)
        if abrir_navegador:
            try:
                webbrowser.open(link)
                print("[AGENTE] Abrindo navegador...")
            except Exception as e:
                print("[AGENTE] Não consegui abrir o navegador:", e)
    else:
        print("[AGENTE] Sua lista está vazia.")

# --------------------- LLM Core ---------------------
def call_llm(prompt, temperature=0.2, max_tokens=400):
    """
    Roteia para OpenAI (se OPENAI_API_KEY) ou Ollama (se OLLAMA_MODEL).
    Saída deve ser uma string JSON seguindo o schema abaixo.
    """
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    provider = os.getenv("LLM_PROVIDER", "auto").lower()
    ollama_model = os.getenv("OLLAMA_MODEL", "llama3.1")
    system = (
        "Você é um parser estrito. Leia o comando de um usuário em PT-BR "
        "sobre lista de compras e RETORNE APENAS um JSON válido com este schema:\n"
        "{\n"
        '  "action": "add" | "remove" | "list" | "send" | "exit",\n'
        '  "items": ["..."],  // array de strings; vazio se não houver\n'
        '  "notes": "string curta opcional"\n'
        "}\n"
        "Regras:\n"
        "- Se o usuário pedir algo como 'bota maçã e 1kg de arroz', action=add e items=['maçã','arroz 1kg'].\n"
        "- Se pedir 'tira o leite' ou 'remove o leite', action=remove e items=['leite'].\n"
        "- Se pedir 'o que tem na lista', 'quais itens', action=list.\n"
        "- Se pedir 'manda pro whatsapp', action=send.\n"
        "- Se disser 'sair', 'valeu', 'pode encerrar', action=exit.\n"
        "- Se ambíguo, tente inferir a intenção mais provável.\n"
        "- SEM TEXTO FORA DO JSON. NÃO COMENTE. NÃO EXPLIQUE."
    )

    fewshot = [
        {
            "role": "user",
            "content": "coloca 2kg de arroz, 1L de leite e três tomates na lista por favor"
        },
        {
            "role": "assistant",
            "content": '{"action":"add","items":["arroz 2kg","leite 1L","tomate 3 unidades"],"notes":""}'
        },
        {
            "role": "user",
            "content": "remove o leite e tira o tomate também"
        },
        {
            "role": "assistant",
            "content": '{"action":"remove","items":["leite","tomate"],"notes":""}'
        },
        {
            "role": "user",
            "content": "o que tem na lista?"
        },
        {
            "role": "assistant",
            "content": '{"action":"list","items":[],"notes":""}'
        },
        {
            "role": "user",
            "content": "pode mandar pro zap"
        },
        {
            "role": "assistant",
            "content": '{"action":"send","items":[],"notes":""}'
        },
        {
            "role": "user",
            "content": "valeu por hoje, tchau"
        },
        {
            "role": "assistant",
            "content": '{"action":"exit","items":[],"notes":""}'
        },
    ]

    # Decide provider
    use_openai = bool(api_key) and (provider in ("auto", "openai"))
    use_ollama = (not use_openai) or (provider == "ollama")

    if use_openai:
        # OpenAI Chat Completions (Responses API equivalents can be adapted)
        try:
            import openai  # lazy import
            client = openai.OpenAI(api_key=api_key)
            resp = client.chat.completions.create(
                model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
                temperature=temperature,
                max_tokens=max_tokens,
                messages=[
                    {"role": "system", "content": system},
                    *fewshot,
                    {"role": "user", "content": prompt},
                ],
            )
            return resp.choices[0].message.content.strip()
        except Exception as e:
            # Fallback para Ollama
            print("[LLM] OpenAI falhou, tentando Ollama:", e)

    # Ollama local
    try:
        r = requests.post(
            "http://localhost:11434/api/chat",
            json={
                "model": ollama_model,
                "messages": [
                    {"role": "system", "content": system},
                    *fewshot,
                    {"role": "user", "content": prompt},
                ],
                "options": {"temperature": temperature, "num_predict": max_tokens},
                "stream": False,
            },
            timeout=60,
        )
        r.raise_for_status()
        data = r.json()
        # Ollama retorna {message: {content: "..."}}
        content = data.get("message", {}).get("content", "").strip()
        return content
    except Exception as e:
        raise RuntimeError(f"Falha ao chamar LLM (Ollama): {e}")

def interpretar_comando_llm(texto):
    """
    Envia o texto para a LLM e retorna um dicionário normalizado:
    {action:str, items:list[str], notes:str}
    Com tratamento robusto de JSON.
    """
    raw = call_llm(texto)
    # Sanitização mínima: pegar o primeiro bloco { ... }
    start = raw.find("{")
    end = raw.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError(f"LLM não retornou JSON válido: {raw}")
    snippet = raw[start:end+1]

    try:
        data = json.loads(snippet)
        action = str(data.get("action", "")).lower().strip()
        items = data.get("items", [])
        if not isinstance(items, list):
            items = []
        items = [str(x).strip() for x in items if str(x).strip()]
        notes = str(data.get("notes", ""))
        return {"action": action, "items": items, "notes": notes}
    except Exception as e:
        raise ValueError(f"Falha parseando JSON da LLM: {e} | Conteúdo: {raw}")

# --------------------- Fluxo de comandos ---------------------
def executar_acao(parsed):
    action = parsed.get("action")
    items = parsed.get("items", [])

    if action == "add":
        if not items:
            print("[AGENTE] O que devo adicionar?")
            return True
        for it in items:
            adicionar_produto(it)
        return True

    elif action == "remove":
        if not items:
            print("[AGENTE] O que devo remover?")
            return True
        for it in items:
            remover_produto(it)
        return True

    elif action == "list":
        listar_produtos()
        return True

    elif action == "send":
        enviar_lista(abrir_navegador=True)
        return True

    elif action == "exit":
        print("[AGENTE] Encerrando...")
        return False

    else:
        print("[AGENTE] Não entendi. Pode reformular?")
        return True

# --------------------- Audio (opcional, igual ao seu) ---------------------
def ouvir_comando_vosk(segundos=5):
    if not VOSK_OK:
        print("[AGENTE] Vosk/sounddevice não disponível. Digite o comando:")
        return input("> ").strip()

    try:
        model = Model("vosk-model")  # baixar PT-BR
        rec = KaldiRecognizer(model, 16000)
        print(f"\n[AGENTE] Fale seu comando ({segundos}s):")
        audio = sd.rec(int(segundos * 16000), samplerate=16000, channels=1, dtype='int16')
        sd.wait()
        rec.AcceptWaveform(audio.tobytes())
        result = json.loads(rec.Result())
        comando = result.get("text", "").strip()
        print(f"[Você]: {comando}")
        return comando
    except Exception as e:
        print("[AGENTE] Falha no áudio:", e)
        return input("> ").strip()

# --------------------- Main ---------------------
def main():
    inicializar_planilha()
    print("=== Agente de Compras com LLM (PT-BR) ===")
    print("Dica: fale/digite livremente. Exemplos:")
    print("- 'coloca 2kg de arroz e três tomates'")
    print("- 'tira o leite'")
    print("- 'o que tem aí?'  |  'manda no zap'  |  'pode encerrar'")

    ativo = True
    while ativo:
        texto = ouvir_comando_vosk(segundos=5)
        if not texto:
            continue
        try:
            parsed = interpretar_comando_llm(texto)
            # print("[DEBUG]", parsed)
            ativo = executar_acao(parsed)
        except Exception as e:
            print("[AGENTE] Erro interpretando com LLM:", e)

if __name__ == "__main__":
    main()
